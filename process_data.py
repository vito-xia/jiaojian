"""交件超时看板离线数据处理工具。

读取项目“数据源”目录下的六类 Excel，生成可直接双击打开的静态看板数据。
运行：python process_data.py [--as-of 2026-07-26]
"""
from __future__ import annotations

import argparse
import json
import posixpath
import re
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any
from xml.etree import ElementTree
from zipfile import ZipFile

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

PLATFORM_ALIAS = {"淘天": "淘宝"}
PLATFORMS = ("抖音", "淘宝", "京东", "快手")
SCORE_SCENES = {"物流停滞-揽收端", "物流停滞-全链路"}
DELIVERY_SCORE_SCENES = {"物流停滞-派送端"}
CONTROL_ACTIONS = {"揽收能力预警", "限制面单新签", "限制面单取号"}
ACTION_SEVERITY = {"揽收能力预警": 1, "限制面单新签": 2, "限制面单取号": 3}
DELIVERY_CONTROL_ACTIONS = {"派送能力预警", "限制面单新签", "限制面单到达"}
DELIVERY_ACTION_SEVERITY = {"派送能力预警": 1, "限制面单新签": 2, "限制面单到达": 3}
EXCLUDED_CUSTOMER_KEYWORDS = ("温宿韵通达", "新疆", "北亩")
EMPTY_DETAIL_VALUES = {"", "-", "--", "—", "/", "无", "暂无"}


def text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"\s+", " ", str(value)).strip()


def number(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(str(value).replace(",", "").replace("%", ""))
    except (TypeError, ValueError):
        return 0.0


def percentage_points(value: Any) -> float:
    amount = number(value)
    if isinstance(value, (int, float)) and 0 < abs(amount) <= 1:
        amount *= 100
    return round(amount, 4)

def integer(value: Any) -> int:
    return int(round(number(value)))


def shipment_interval(value: int | None) -> str:
    if value is None:
        return "无法计算"
    if value <= 50:
        return "50以内"
    if value <= 100:
        return "50-100"
    if value <= 300:
        return "100-300"
    if value <= 500:
        return "300-500"
    if value <= 1000:
        return "500-1000"
    if value <= 5000:
        return "1000-5000"
    return "5000+"


def jd_delivery_metrics(timeout_36h: int, timeout_rate_36h: float, timeout_48h: int) -> dict[str, Any]:
    """京东派生口径：源表超时率为百分点，计算时需先除以 100。"""
    rate = number(timeout_rate_36h)
    shipment_volume = int(round(timeout_36h * 100 / rate)) if rate > 0 else None
    timeout_rate_48h = round(timeout_48h / shipment_volume * 100, 4) if shipment_volume else None
    return {
        "shipment_volume": shipment_volume,
        "shipment_interval": shipment_interval(shipment_volume),
        "timeout_rate_48h": timeout_rate_48h,
    }


def customer_is_excluded(customer_name: Any) -> bool:
    name = text(customer_name)
    return any(keyword in name for keyword in EXCLUDED_CUSTOMER_KEYWORDS)


def has_detail(value: Any) -> bool:
    return text(value) not in EMPTY_DETAIL_VALUES


def excel_column_number(letters: str) -> int:
    result = 0
    for character in letters.upper():
        result = result * 26 + ord(character) - ord("A") + 1
    return result


def merged_reason_action_rows(path: Path, sheet_name: str) -> set[int]:
    main_namespace = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    office_relationships = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    package_relationships = "http://schemas.openxmlformats.org/package/2006/relationships"
    with ZipFile(path) as archive:
        workbook_root = ElementTree.fromstring(archive.read("xl/workbook.xml"))
        relationships_root = ElementTree.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        sheet = next(
            node for node in workbook_root.findall(f".//{{{main_namespace}}}sheet")
            if node.attrib.get("name") == sheet_name
        )
        relationship_id = sheet.attrib[f"{{{office_relationships}}}id"]
        relationship = next(
            node for node in relationships_root.findall(f".//{{{package_relationships}}}Relationship")
            if node.attrib.get("Id") == relationship_id
        )
        target = relationship.attrib["Target"]
        sheet_path = target.lstrip("/") if target.startswith("/") else posixpath.normpath(f"xl/{target}")
        sheet_root = ElementTree.fromstring(archive.read(sheet_path))

    merged_rows: set[int] = set()
    for cell_range in sheet_root.findall(f".//{{{main_namespace}}}mergeCell"):
        reference = cell_range.attrib.get("ref", "")
        match = re.fullmatch(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", reference)
        if not match:
            continue
        start_column = excel_column_number(match.group(1))
        end_column = excel_column_number(match.group(3))
        if start_column > 20 or end_column < 21:
            continue
        merged_rows.update(range(int(match.group(2)), int(match.group(4)) + 1))
    return merged_rows

def date_text(value: Any, epoch=None, default_year: int | None = None) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        try:
            return from_excel(value, epoch).date().isoformat()
        except (TypeError, ValueError, OverflowError):
            return ""
    raw = text(value)
    if re.fullmatch(r"\d+(\.0)?", raw) and epoch is not None:
        try:
            return from_excel(float(raw), epoch).date().isoformat()
        except (TypeError, ValueError, OverflowError):
            pass
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(raw[:10], fmt).date().isoformat()
        except ValueError:
            pass
    match = re.search(r"(?:(\d{4})年)?(\d{1,2})月(\d{1,2})日", raw)
    if match and (match.group(1) or default_year):
        return date(int(match.group(1) if match.group(1) else default_year), int(match.group(2)), int(match.group(3))).isoformat()
    return ""


def iso_day(value: str) -> date:
    return datetime.strptime(value, "%Y-%m-%d").date()


def write_json(path: Path, payload: Any, pretty: bool = False) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    with tmp.open("w", encoding="utf-8", newline="\n") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2 if pretty else None, separators=None if pretty else (",", ":"))
        handle.write("\n")
    tmp.replace(path)


def locate(data_dir: Path, prefix: str) -> Path:
    matches = sorted(data_dir.glob(f"{prefix}*.xlsx"))
    if len(matches) != 1:
        raise FileNotFoundError(f"期望找到 1 个 {prefix}*.xlsx，实际找到 {len(matches)} 个")
    return matches[0]


def read_timeout(data_dir: Path, year: int) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    platform_pattern = "|".join(re.escape(platform) for platform in PLATFORMS)
    for path in sorted((data_dir / "①交件超时").glob("*.xlsx")):
        match = re.match(rf"({platform_pattern})_(\d{{1,2}})月(\d{{1,2}})日\.xlsx$", path.name)
        if not match or path.name.startswith("~$"):
            continue
        platform, month, day = match.groups()
        record_date = date(year, int(month), int(day)).isoformat()
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=3, values_only=True):
            if not row or text(row[1] if len(row) > 1 else "") in ("", "合计"):
                continue
            record = {
                "platform": platform,
                "date": record_date,
                "branch": text(row[1]),
                "customer": text(row[2]),
                "customer_code": text(row[3]),
                "has_shipping_fallback": text(row[4]),
                "has_history_no_goods": text(row[5]),
                "timeout_24h": integer(row[6]),
                "timeout_36h": integer(row[7]),
                "timeout_rate_36h": round(number(row[8]), 4),
                "timeout_48h": integer(row[9]),
                "timeout_72h": integer(row[10]),
                "timeout_96h": integer(row[11]),
                "timeout_120h": integer(row[12]),
            }
            if platform == "京东":
                record.update(jd_delivery_metrics(record["timeout_36h"], record["timeout_rate_36h"], record["timeout_48h"]))
            records.append(record)
        workbook.close()
    return records


def read_top5(data_dir: Path, year: int) -> tuple[list[dict[str, Any]], list[dict[str, Any]], int]:
    path = locate(data_dir, "②")
    merged_tu_rows = merged_reason_action_rows(path, "TOP客户累计改善清单")
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["TOP客户累计改善清单"]
    records, branch_top5_rows, bad_dates = [], [], 0
    platform_columns = {
        "抖音": {"timeout_24h": 11, "timeout_36h": 12, "timeout_rate_36h": 13, "timeout_48h": 14},
        "淘宝": {"timeout_24h": 15, "timeout_36h": 16, "timeout_rate_36h": 17, "timeout_48h": 18},
    }
    for row_number, row in enumerate(sheet.iter_rows(min_row=10, values_only=True), start=10):
        if not row or row[1] in (None, ""):
            continue
        parsed_date = date_text(row[1], workbook.epoch, year)
        if not parsed_date:
            bad_dates += 1
            continue
        source_platform = PLATFORM_ALIAS.get(text(row[9]), text(row[9]))
        reason = text(row[19])
        action = text(row[20])
        base = {
            "date": parsed_date,
            "branch": text(row[3]),
            "customer": text(row[5]),
            "customer_code": text(row[6]),
            "has_shipping_fallback": text(row[7]),
            "source_platform": source_platform,
            "ranking_count": integer(row[10]),
        }
        if source_platform in platform_columns:
            columns = platform_columns[source_platform]
            branch_top5_rows.append({
                **base,
                "source_row": row_number,
                "platform": source_platform,
                "timeout_36h": integer(row[columns["timeout_36h"]]),
                "timeout_rate_36h": percentage_points(row[columns["timeout_rate_36h"]]),
                "reason": reason,
                "action": action,
                "feedback_merged": row_number in merged_tu_rows,
            })
        for platform, columns in platform_columns.items():
            timeout_36h = integer(row[columns["timeout_36h"]])
            if timeout_36h <= 0:
                continue
            records.append({
                **base,
                "platform": platform,
                "timeout_24h": integer(row[columns["timeout_24h"]]),
                "timeout_36h": timeout_36h,
                "timeout_rate_36h": percentage_points(row[columns["timeout_rate_36h"]]),
                "timeout_48h": integer(row[columns["timeout_48h"]]),
            })
    workbook.close()
    return records, branch_top5_rows, bad_dates

def read_mapping(data_dir: Path) -> dict[str, dict[str, str]]:
    path = locate(data_dir, "③")
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    mapping = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        branch = text(row[2] if len(row) > 2 else "")
        if not branch:
            continue
        parent = text(row[7] if len(row) > 7 else "") or branch
        mapping[branch] = {
            "code": text(row[1]), "branch": branch, "short_name": text(row[3]),
            "nature": text(row[4]), "parent_code": text(row[6]), "parent_name": parent,
            "region": text(row[9]), "province": text(row[11]),
        }
    workbook.close()
    return mapping


def parse_cumulative_score(raw: Any) -> float:
    return parse_scene_scores(raw, SCORE_SCENES)

def parse_scene_scores(raw: Any, scenes: set[str]) -> float:
    value = text(raw)
    total = 0.0
    for scene in scenes:
        match = re.search(re.escape(scene) + r"[：:]\s*(-?\d+(?:\.\d+)?)", value)
        if match:
            total += float(match.group(1))
    return total


def read_scores(data_dir: Path) -> tuple[list[dict[str, Any]], dict[str, dict[str, float]], dict[str, dict[str, float]]]:
    path = locate(data_dir, "④")
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    records = []
    daily: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    cumulative: dict[str, dict[str, float]] = defaultdict(dict)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        branch, scene = text(row[0]), text(row[9])
        violation_date = date_text(row[16], workbook.epoch)
        if not branch or not violation_date:
            continue
        current = number(row[14])
        snapshot = parse_cumulative_score(row[15])
        records.append({
            "branch": branch,
            "scene": scene,
            "date": violation_date,
            "current_score": current,
            "cumulative_stagnant_score": snapshot,
            "shipment_timeout_rate": percentage_points(row[20]) if has_detail(row[20]) else None,
            "shipment_timeout_abnormal_count": integer(row[21]) if has_detail(row[21]) else None,
            "shipment_timeout_operation_count": integer(row[22]) if has_detail(row[22]) else None,
        })
        if scene in SCORE_SCENES:
            daily[branch][violation_date] += current
            cumulative[branch][violation_date] = max(cumulative[branch].get(violation_date, 0), snapshot)
    workbook.close()
    return records, {b: dict(v) for b, v in daily.items()}, {b: dict(v) for b, v in cumulative.items()}


def read_controls(data_dir: Path) -> list[dict[str, Any]]:
    path = locate(data_dir, "⑤")
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    records = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        merchant_id = text(row[5])
        records.append({
            "control_id": text(row[0]), "branch_code": text(row[1]), "branch": text(row[2]),
            "merchant_id": merchant_id, "merchant_name": text(row[6]), "violation_scene": text(row[7]),
            "control_status": text(row[8]), "control_action": text(row[9]),
            "start_date": date_text(row[10], workbook.epoch), "end_date": date_text(row[11], workbook.epoch),
            "control_mechanism": text(row[12]), "control_category": text(row[13]),
            "is_branch_level": not bool(merchant_id), "is_merchant_level": bool(merchant_id),
        })
    workbook.close()
    return records



def read_delivery_monitor(data_dir: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, dict[str, float]], dict[str, dict[str, float]], list[str]]:
    path = locate(data_dir, "⑥")
    workbook = load_workbook(path, read_only=True, data_only=True)

    control_sheet = workbook["每日派送管控"]
    controls = []
    for row in control_sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        merchant_id = text(row[5])
        controls.append({
            "control_id": text(row[0]), "branch_code": text(row[1]), "branch": text(row[2]),
            "carrier": text(row[3]), "region": text(row[4]), "merchant_id": merchant_id,
            "merchant_name": text(row[6]), "violation_scene": text(row[7]),
            "control_status": text(row[8]), "control_action": text(row[9]),
            "start_date": date_text(row[10], workbook.epoch), "end_date": date_text(row[11], workbook.epoch),
            "control_mechanism": text(row[12]), "control_category": text(row[13]),
            "is_branch_level": not bool(merchant_id), "is_merchant_level": bool(merchant_id),
        })

    score_sheet = workbook["每日派送积分"]
    score_headers = [text(cell.value) for cell in score_sheet[1]]
    score_rows = []
    daily: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    cumulative: dict[str, dict[str, float]] = defaultdict(dict)
    for row in score_sheet.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 17:
            continue
        branch, violation_date = text(row[0]), date_text(row[16], workbook.epoch)
        if not branch or not violation_date:
            continue
        scene = text(row[9])
        current = number(row[14])
        snapshot = parse_scene_scores(row[15], DELIVERY_SCORE_SCENES)
        detail = {
            "branch": branch, "branch_code": text(row[1]), "province": text(row[2]),
            "city": text(row[3]), "district": text(row[4]), "street": text(row[5]),
            "carrier": text(row[6]), "merchant_id": text(row[7]), "merchant_name": text(row[8]),
            "scene": scene, "abnormal_level": text(row[10]), "collaboration_status": text(row[11]),
            "reason_level_1": text(row[12]), "reason_level_2": text(row[13]),
            "current_score": current, "cumulative_score": snapshot, "date": violation_date,
            "delivery_signature_timeout": text(row[38]), "delivery_signature_timeout_abnormal_count": integer(row[39]),
            "delivery_signature_timeout_operation_count": integer(row[40]), "not_dispatched_timeout": text(row[41]),
            "not_dispatched_timeout_abnormal_count": integer(row[42]), "not_dispatched_timeout_operation_count": integer(row[43]),
        }
        score_rows.append(detail)
        if scene in DELIVERY_SCORE_SCENES:
            daily[branch][violation_date] += current
            cumulative[branch][violation_date] = max(cumulative[branch].get(violation_date, 0), snapshot)
    workbook.close()
    return controls, score_rows, {b: dict(v) for b, v in daily.items()}, {b: dict(v) for b, v in cumulative.items()}, score_headers
def parent_of(branch: str, mapping: dict[str, dict[str, str]]) -> str:
    return mapping.get(branch, {}).get("parent_name") or branch


def build_shortage_history(top5: list[dict[str, Any]], mapping: dict[str, dict[str, str]]) -> dict[str, dict[str, Any]]:
    work: dict[tuple[str, str], dict[str, Any]] = defaultdict(lambda: {"months": defaultdict(set), "branches": set(), "customers": set()})
    for row in top5:
        parent = parent_of(row["branch"], mapping)
        bucket = work[(row["platform"], parent)]
        bucket["months"][row["date"][:7]].add(row["date"])
        bucket["branches"].add(row["branch"])
        bucket["customers"].add(row["customer"])
    result: dict[str, dict[str, Any]] = {p: {} for p in PLATFORMS}
    for (platform, parent), bucket in work.items():
        result[platform][parent] = {
            "months": [{"month": month, "days": len(days)} for month, days in sorted(bucket["months"].items())],
            "branches": sorted(bucket["branches"]), "customer_count": len(bucket["customers"]),
        }
    return result


def build_shortage_history_all(top5: list[dict[str, Any]], mapping: dict[str, dict[str, str]]) -> dict[str, dict[str, Any]]:
    work: dict[str, dict[str, Any]] = defaultdict(lambda: {"months": defaultdict(set), "branches": set(), "customers": set()})
    for row in top5:
        parent = parent_of(row["branch"], mapping)
        bucket = work[parent]
        bucket["months"][row["date"][:7]].add(row["date"])
        bucket["branches"].add(row["branch"])
        bucket["customers"].add(row["customer"])
    return {
        parent: {
            "months": [{"month": month, "days": len(days)} for month, days in sorted(bucket["months"].items())],
            "branches": sorted(bucket["branches"]),
            "customer_count": len(bucket["customers"]),
        }
        for parent, bucket in work.items()
    }


def build_branch_top5_data(branch_top5_rows: list[dict[str, Any]]) -> dict[str, dict[str, list[dict[str, Any]]]]:
    result: dict[str, dict[str, list[dict[str, Any]]]] = {
        "抖音": defaultdict(list),
        "淘宝": defaultdict(list),
    }
    for row in branch_top5_rows:
        platform = row["platform"]
        branch = row["branch"]
        if platform in result and branch:
            result[platform][branch].append(row)
    return {
        platform: {
            branch: sorted(rows, key=lambda row: (row["date"], row["source_row"]), reverse=True)
            for branch, rows in sorted(by_branch.items())
        }
        for platform, by_branch in result.items()
    }

def clearout_type(mechanism: str) -> str:
    if "熔断制" in mechanism:
        return "熔断制"
    if "积分制" in mechanism:
        return "积分制"
    return mechanism or "未知"


def build_control_index(controls: list[dict[str, Any]], mapping: dict[str, dict[str, str]], as_of: str):
    relevant = [r for r in controls if not r["start_date"] or r["start_date"] <= as_of]
    current_by_branch: dict[str, list[dict[str, Any]]] = defaultdict(list)
    merchant_executing = Counter()
    clearouts = []
    for row in relevant:
        executing = "执行中" in row["control_status"]
        if row["is_branch_level"] and executing and row["control_action"] in CONTROL_ACTIONS:
            current_by_branch[row["branch"]].append(row)
        if row["is_merchant_level"] and executing:
            merchant_executing[row["branch"]] += 1
        if row["is_branch_level"] and row["control_action"] == "限制面单取号" and ("执行中" in row["control_status"] or row["control_status"] == "已完结"):
            clearouts.append(row)
    current = {}
    for branch, rows in current_by_branch.items():
        best = max(rows, key=lambda r: (ACTION_SEVERITY.get(r["control_action"], 0), r["start_date"]))
        current[branch] = {"action": best["control_action"], "status": best["control_status"], "start_date": best["start_date"]}
    parent_rows: dict[str, list[dict[str, Any]]] = defaultdict(list)
    branch_counts = Counter()
    for row in clearouts:
        parent_rows[parent_of(row["branch"], mapping)].append(row)
        branch_counts[row["branch"]] += 1
    parent_info = {}
    for parent, rows in parent_rows.items():
        latest = max(rows, key=lambda r: (r["start_date"], r["control_id"]))
        parent_info[parent] = {"count": len(rows), "last_date": latest["start_date"], "last_type": clearout_type(latest["control_mechanism"])}
    return current, dict(merchant_executing), parent_info, dict(branch_counts), clearouts


def build_score_calculator(daily_scores: dict[str, dict[str, float]], window_days: int = 16):
    cache: dict[tuple[str, str], float] = {}
    def rolling(branch: str, end_day: str) -> float:
        key = (branch, end_day)
        if key not in cache:
            end = iso_day(end_day)
            start = end - timedelta(days=window_days - 1)
            cache[key] = sum(value for day, value in daily_scores.get(branch, {}).items() if start <= iso_day(day) <= end)
        return round(cache[key], 2)
    return rolling



def build_delivery_monitor(controls, score_rows, daily_scores, cumulative_scores, mapping):
    score_dates = sorted({row["date"] for row in score_rows if row["scene"] in DELIVERY_SCORE_SCENES and row["date"]})
    control_dates = sorted({row["start_date"] for row in controls if row["start_date"]})
    as_of = score_dates[-1] if score_dates else (control_dates[-1] if control_dates else "")
    control_as_of = control_dates[-1] if control_dates else as_of
    rolling_score = build_score_calculator(daily_scores)
    control_by_date, high_scores_by_date, score_details_by_date = {}, {}, {}
    control_days = sorted(set(score_dates) | ({control_as_of} if control_as_of else set()))
    branch_controls = [row for row in controls if row["is_branch_level"] and row["control_action"] in DELIVERY_CONTROL_ACTIONS and row["start_date"]]
    for day in control_days:
        end = iso_day(day)
        start = end - timedelta(days=6)
        page = []
        for row in branch_controls:
            event_day = iso_day(row["start_date"])
            if not start <= event_day <= end:
                continue
            page.append({
                "control_id": row["control_id"], "date": row["start_date"], "branch_code": row["branch_code"],
                "branch": row["branch"], "parent_name": parent_of(row["branch"], mapping), "region": row["region"],
                "violation_scene": row["violation_scene"], "control_status": row["control_status"],
                "control_action": row["control_action"], "rolling_score": rolling_score(row["branch"], day), "end_date": row["end_date"],
                "control_mechanism": row["control_mechanism"], "control_category": row["control_category"],
            })
        page.sort(key=lambda row: (row["date"], DELIVERY_ACTION_SEVERITY.get(row["control_action"], 0), row["branch"]), reverse=True)
        control_by_date[day] = page

        high = []
        for branch in daily_scores:
            score = rolling_score(branch, day)
            if score < 12:
                continue
            dates = sorted(date for date in daily_scores[branch] if date <= day)
            latest_date = dates[-1] if dates else ""
            latest_rows = [row for row in score_rows if row["branch"] == branch and row["date"] == latest_date and row["scene"] in DELIVERY_SCORE_SCENES]
            latest_row = max(latest_rows, key=lambda row: (row["current_score"], row["cumulative_score"])) if latest_rows else {}
            high.append({
                "branch": branch, "branch_code": latest_row.get("branch_code", ""), "parent_name": parent_of(branch, mapping),
                "stagnant_score": score, "latest_daily_score": daily_scores[branch].get(latest_date, 0),
                "latest_score_date": latest_date, "latest_abnormal_level": latest_row.get("abnormal_level", ""),
                "latest_cumulative_score": cumulative_scores.get(branch, {}).get(latest_date, 0),
            })
        high.sort(key=lambda row: (-row["stagnant_score"], -row["latest_daily_score"], row["branch"]))
        high_scores_by_date[day] = high
        score_details_by_date[day] = [row for row in score_rows if row["date"] == day and row["scene"] in DELIVERY_SCORE_SCENES]

    trends = {}
    for branch, days in daily_scores.items():
        trends[branch] = {
            "parent_name": parent_of(branch, mapping),
            "series": [{"date": day, "daily_score": round(days.get(day, 0), 2), "rolling_score": rolling_score(branch, day)} for day in score_dates],
        }
    return {
        "as_of": as_of, "control_as_of": control_as_of, "score_dates": score_dates,
        "control_dates": control_days, "controls_by_date": control_by_date,
        "high_scores_by_date": high_scores_by_date, "score_details_by_date": score_details_by_date,
        "trends": trends,
    }
def build_trends(timeout_rows: list[dict[str, Any]], mapping: dict[str, dict[str, str]]) -> dict[str, dict[str, Any]]:
    work: dict[str, dict[str, dict[str, Any]]] = {p: defaultdict(dict) for p in PLATFORMS}
    for row in timeout_rows:
        branch_box = work[row["platform"]][row["branch"]]
        if row["platform"] == "京东":
            customer_key = ("code", row["customer_code"]) if row["customer_code"] else ("name", row["customer"])
        else:
            customer_key = row["customer"]
        customer_box = branch_box.setdefault(customer_key, {
            "customer": row["customer"], "customer_code": row["customer_code"],
            "has_shipping_fallback": row["has_shipping_fallback"], "series": [], "total_36h": 0,
        })
        point = {k: row[k] for k in ("date", "timeout_24h", "timeout_36h", "timeout_rate_36h", "timeout_48h", "timeout_72h", "timeout_96h", "timeout_120h")}
        if row["platform"] == "京东":
            point.update({k: row.get(k) for k in ("shipment_volume", "shipment_interval", "timeout_rate_48h")})
        customer_box["series"].append(point)
        customer_box["total_36h"] += row["timeout_36h"]
    result = {p: {} for p in PLATFORMS}
    for platform in PLATFORMS:
        for branch, customers in work[platform].items():
            items = list(customers.values())
            for item in items:
                item["series"].sort(key=lambda row: row["date"])
            items.sort(key=lambda item: (-item["total_36h"], item["customer"]))
            result[platform][branch] = {"parent_name": parent_of(branch, mapping), "customers": items}
    return result


def build_branch_score_trends(score_rows: list[dict[str, Any]]) -> dict[str, dict[str, list[dict[str, Any]]]]:
    work: dict[str, dict[str, dict[str, dict[str, Any]]]] = defaultdict(lambda: defaultdict(dict))
    for row in score_rows:
        scene = row.get("scene", "")
        branch = row.get("branch", "")
        day = row.get("date", "")
        if not branch or scene not in SCORE_SCENES or not day:
            continue
        point = work[branch][scene].setdefault(day, {
            "date": day,
            "score": 0.0,
            "shipment_timeout_abnormal_count": 0,
            "_rate_weighted_sum": 0.0,
            "_rate_weight": 0,
            "_rate_values": [],
        })
        point["score"] += number(row.get("current_score"))
        count = row.get("shipment_timeout_abnormal_count")
        if count is not None:
            point["shipment_timeout_abnormal_count"] += number(count)
        rate = row.get("shipment_timeout_rate")
        operation_count = row.get("shipment_timeout_operation_count")
        if rate is not None:
            point["_rate_values"].append(number(rate))
            if operation_count is not None and number(operation_count) > 0:
                point["_rate_weighted_sum"] += number(rate) * number(operation_count)
                point["_rate_weight"] += number(operation_count)

    result: dict[str, dict[str, list[dict[str, Any]]]] = {}
    for branch, scenes in work.items():
        result[branch] = {}
        for scene, points_by_date in scenes.items():
            points = []
            for point in sorted(points_by_date.values(), key=lambda item: item["date"]):
                values = point.pop("_rate_values")
                rate_weight = point.pop("_rate_weight")
                weighted_sum = point.pop("_rate_weighted_sum")
                if rate_weight > 0:
                    point["shipment_timeout_rate"] = round(weighted_sum / rate_weight, 4)
                elif values:
                    point["shipment_timeout_rate"] = round(sum(values) / len(values), 4)
                else:
                    point["shipment_timeout_rate"] = None
                point["score"] = round(point["score"], 2)
                point["shipment_timeout_abnormal_count"] = round(point["shipment_timeout_abnormal_count"])
                points.append(point)
            result[branch][scene] = points
    return result


def build_dashboard(timeout_rows, top5, branch_top5_rows, mapping, score_rows, daily_scores, cumulative_scores, controls, delivery_controls, delivery_score_rows, delivery_daily_scores, delivery_cumulative_scores, as_of: str, bad_top5_dates: int):
    raw_timeout_count = len(timeout_rows)
    raw_top5_count = len(top5)
    timeout_rows = [row for row in timeout_rows if not customer_is_excluded(row["customer"])]
    top5 = [row for row in top5 if not customer_is_excluded(row["customer"])]
    branch_top5_rows = [row for row in branch_top5_rows if not customer_is_excluded(row["customer"])]
    excluded_timeout_count = raw_timeout_count - len(timeout_rows)
    excluded_top5_count = raw_top5_count - len(top5)
    dates_by_platform = {p: sorted({r["date"] for r in timeout_rows if r["platform"] == p}) for p in PLATFORMS}
    shortage = build_shortage_history(top5, mapping)
    shortage_all = build_shortage_history_all(top5, mapping)
    branch_top5_data = build_branch_top5_data(branch_top5_rows)
    score_dates = sorted({r["date"] for r in score_rows})
    control_dates = sorted({r["start_date"] for r in controls if r["start_date"]})
    control_as_of = control_dates[-1] if control_dates else as_of
    # 平台管控是独立的实时快照，允许比交件 T-1 更新；不能用 as_of 截断 7 月 31 日的管控。
    current_controls, merchant_counts, parent_clear, branch_clear_counts, clearouts = build_control_index(controls, mapping, control_as_of)
    rolling_score = build_score_calculator(daily_scores)
    delivery_monitor = build_delivery_monitor(delivery_controls, delivery_score_rows, delivery_daily_scores, delivery_cumulative_scores, mapping)
    top10_by_date = {p: {} for p in PLATFORMS}
    top60_by_date = {"京东": {}}
    for platform in PLATFORMS:
        for day in dates_by_platform[platform]:
            rows_for_day = (r for r in timeout_rows if r["platform"] == platform and r["date"] == day)
            if platform == "京东":
                daily = sorted(rows_for_day, key=lambda r: (-r["timeout_48h"], -number(r.get("timeout_rate_48h")), -r["timeout_36h"], r["branch"]))[:60]
            else:
                daily = sorted(rows_for_day, key=lambda r: (-r["timeout_36h"], -r["timeout_rate_36h"], r["branch"]))[:10]
            enriched = []
            for rank, row in enumerate(daily, 1):
                branch, parent = row["branch"], parent_of(row["branch"], mapping)
                clear = parent_clear.get(parent, {"count": 0, "last_date": "", "last_type": ""})
                ctrl = current_controls.get(branch, {"action": "", "status": "", "start_date": ""})
                history_shortage = shortage.get(platform, {}).get(parent)
                if history_shortage is None and platform not in ("抖音", "淘宝"):
                    history_shortage = shortage_all.get(parent)
                enriched.append({**row, "rank": rank, "parent_name": parent,
                    "stagnant_score": rolling_score(branch, day) if platform == "抖音" else None,
                    "current_control": ctrl["action"] if platform == "抖音" else "",
                    "merchant_control_count": merchant_counts.get(branch, 0) if platform == "抖音" else None,
                    "branch_clearout_count": branch_clear_counts.get(branch, 0) if platform == "抖音" else None,
                    "clearout_count": clear["count"] if platform == "抖音" else None,
                    "last_clearout_date": clear["last_date"] if platform == "抖音" else "",
                    "last_clearout_type": clear["last_type"] if platform == "抖音" else "",
                    "history_shortage": history_shortage or {"months": [], "branches": [], "customer_count": 0},
                })
            if platform == "京东":
                top60_by_date[platform][day] = enriched
                top10_by_date[platform][day] = enriched[:10]
            else:
                top10_by_date[platform][day] = enriched
    branch_events = [r for r in controls if r["is_branch_level"] and r["control_action"] in CONTROL_ACTIONS and r["start_date"]]
    control_by_date, high_scores_by_date = {}, {}
    # 保留交件日期索引，同时补充最新管控快照日期，供前端在选择 7 月 30 日时展示 7 月 31 日管控。
    control_days = sorted(set(dates_by_platform["抖音"]) | {control_as_of})
    for day in control_days:
        end, start = iso_day(day), iso_day(day) - timedelta(days=6)
        page = []
        for row in branch_events:
            event_day = iso_day(row["start_date"])
            if not start <= event_day <= end:
                continue
            parent = parent_of(row["branch"], mapping)
            clear = parent_clear.get(parent, {"count": 0, "last_date": "", "last_type": ""})
            page.append({"date": row["start_date"], "branch": row["branch"], "parent_name": parent,
                "control_action": row["control_action"], "control_status": row["control_status"],
                "stagnant_score": rolling_score(row["branch"], day), "clearout_count": clear["count"],
                "last_clearout_date": clear["last_date"], "last_clearout_type": clear["last_type"]})
        page.sort(key=lambda r: (r["date"], ACTION_SEVERITY.get(r["control_action"], 0), r["branch"]), reverse=True)
        control_by_date[day] = page
        high = []
        for branch in daily_scores:
            score = rolling_score(branch, day)
            if score < 6:
                continue
            parent = parent_of(branch, mapping)
            clear = parent_clear.get(parent, {"count": 0, "last_date": "", "last_type": ""})
            high.append({"branch": branch, "parent_name": parent, "stagnant_score": score,
                "clearout_count": clear["count"], "last_clearout_date": clear["last_date"], "last_clearout_type": clear["last_type"]})
        high.sort(key=lambda r: (-r["stagnant_score"], -r["clearout_count"], r["branch"]))
        high_scores_by_date[day] = high
    all_branches = {r["branch"] for r in timeout_rows + top5 + score_rows + controls if r.get("branch")}
    unmatched = sorted(branch for branch in all_branches if branch not in mapping)
    timeout_dates = sorted({r["date"] for r in timeout_rows})
    current_branch_executing = len({b for b in current_controls})
    current_merchant_executing = sum(merchant_counts.values())
    example_parent = "广东佛山南海新河村公司"
    platform_payloads = {p: {"dates": dates_by_platform[p], "top10_by_date": top10_by_date[p]} for p in PLATFORMS}
    platform_payloads["京东"]["top60_by_date"] = top60_by_date["京东"]
    return {
        "meta": {
            "as_of": as_of, "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "score_as_of": score_dates[-1] if score_dates else "", "score_window_start": (iso_day(as_of) - timedelta(days=15)).isoformat(),
            "control_as_of": control_as_of, "control_window_start": (iso_day(control_as_of) - timedelta(days=6)).isoformat() if control_as_of else "", "timeout_start": timeout_dates[0] if timeout_dates else "",
            "supported_platforms": list(PLATFORMS),
            "source_rows": {"timeout": raw_timeout_count, "top5": raw_top5_count, "mapping": len(mapping), "scores": len(score_rows), "controls": len(controls)},
            "analysis_rows": {"timeout": len(timeout_rows), "top5": len(top5)},
            "quality": {
                "unmatched_branch_count": len(unmatched),
                "unmatched_branch_sample": unmatched[:20],
                "invalid_top5_dates": bad_top5_dates,
                "excluded_customer_keywords": list(EXCLUDED_CUSTOMER_KEYWORDS),
                "excluded_timeout_rows": excluded_timeout_count,
                "excluded_top5_rows": excluded_top5_count,
            },
            "summary": {"executing_branch_controls": current_branch_executing, "executing_merchant_controls": current_merchant_executing, "historical_clearouts": len(clearouts)},
            "example_check": shortage.get("抖音", {}).get(example_parent, {"months": [], "branches": []}),
        },
        "platforms": platform_payloads,
        "controls_by_date": control_by_date, "high_scores_by_date": high_scores_by_date,
        "history_lookup": shortage,
        "history_all_lookup": shortage_all,
        "branch_top5_data": branch_top5_data,
        "branch_score_trends": build_branch_score_trends(score_rows),
        "trends": build_trends(timeout_rows, mapping),
        "delivery_monitor": delivery_monitor,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="生成交件超时静态看板数据")
    parser.add_argument("--data-dir", type=Path, default=Path(__file__).resolve().parent / "数据源")
    parser.add_argument("--output-dir", type=Path, default=Path(__file__).resolve().parent / "data")
    parser.add_argument("--as-of", help="看板 T-1 日期，默认取交件数据最大日期")
    parser.add_argument("--year", type=int, default=datetime.now().year)
    args = parser.parse_args()
    print("读取 6 类数据源…")
    timeout_rows = read_timeout(args.data_dir, args.year)
    if not timeout_rows:
        raise RuntimeError("未读取到交件超时数据")
    top5, branch_top5_rows, bad_dates = read_top5(args.data_dir, args.year)
    mapping = read_mapping(args.data_dir)
    score_rows, daily_scores, cumulative_scores = read_scores(args.data_dir)
    controls = read_controls(args.data_dir)
    delivery_controls, delivery_score_rows, delivery_daily_scores, delivery_cumulative_scores, delivery_score_headers = read_delivery_monitor(args.data_dir)
    as_of = args.as_of or max(row["date"] for row in timeout_rows)
    if as_of not in {row["date"] for row in timeout_rows}:
        raise ValueError(f"--as-of {as_of} 不在交件数据日期中")
    dashboard = build_dashboard(timeout_rows, top5, branch_top5_rows, mapping, score_rows, daily_scores, cumulative_scores, controls, delivery_controls, delivery_score_rows, delivery_daily_scores, delivery_cumulative_scores, as_of, bad_dates)
    dashboard["delivery_monitor"]["score_headers"] = delivery_score_headers
    output = args.output_dir
    write_json(output / "timeout_daily.json", timeout_rows)
    write_json(output / "top5_control.json", top5)
    write_json(output / "branch_mapping.json", mapping)
    write_json(output / "platform_scores.json", score_rows)
    write_json(output / "platform_control.json", controls)
    write_json(output / "delivery_control.json", delivery_controls)
    write_json(output / "delivery_score.json", delivery_score_rows)
    write_json(output / "dashboard_data.json", dashboard)
    write_json(output / "data_quality_report.json", dashboard["meta"])
    bundle = "window.__JIAOJIAN_DASHBOARD__=" + json.dumps(dashboard, ensure_ascii=False, separators=(",", ":")).replace("</", "<\\/") + ";\n"
    (output / "dashboard_bundle.js").write_text(bundle, encoding="utf-8", newline="\n")
    print(f"完成：T-1={as_of}，交件 {len(timeout_rows)} 条，TOP5 {len(top5)} 条，交件积分 {len(score_rows)} 条，交件管控 {len(controls)} 条，派送积分 {len(delivery_score_rows)} 条，派送管控 {len(delivery_controls)} 条")
    print(f"输出：{output.resolve()}")
    check = dashboard["meta"]["example_check"]
    print("示例核验：广东佛山南海新河村公司 / 抖音", check.get("months", []), check.get("branches", []))


if __name__ == "__main__":
    main()